import sys
import os
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from openpyxl import load_workbook
from app.database import SessionLocal
from app.models import (
    Product, DisplaySpecs, HardwareSpecs,
    PhoneSpecs, CameraSpecs
)


def parse_bool(value) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        return value.strip().upper() == "TRUE"
    return False


def parse_num(value):
    if value is None or value == "":
        return None
    try:
        return float(value)
    except:
        return None


def parse_int(value):
    if value is None or value == "":
        return None
    try:
        return int(float(value))
    except:
        return None


def parse_str(value):
    if value is None:
        return None
    v = str(value).strip()
    return v if v else None


def import_products(filepath: str):
    if not os.path.exists(filepath):
        print(f"❌ File not found: {filepath}")
        return

    wb = load_workbook(filepath, data_only=True)
    ws = wb["Products"]

    headers = [cell.value for cell in ws[2]]
    rows = list(ws.iter_rows(min_row=4, values_only=True))

    db = SessionLocal()
    success = 0
    errors = 0

    for row_idx, row in enumerate(rows, start=4):
        if not any(row):
            continue

        data = dict(zip(headers, row))

        name = parse_str(data.get("name"))
        if not name:
            continue

        slug = parse_str(data.get("slug"))
        if not slug:
            print(f"⚠️  Row {row_idx} — '{name}' skipped: missing slug")
            errors += 1
            continue

        existing = db.query(Product).filter(Product.slug == slug).first()
        if existing:
            print(f"⚠️  Row {row_idx} — '{name}' skipped: slug '{slug}' already exists")
            errors += 1
            continue

        try:
            product = Product(
                name=name,
                brand=parse_str(data.get("brand")) or "Unknown",
                model=parse_str(data.get("model")),
                category=parse_str(data.get("category")) or "phone",
                slug=slug,
                variant_group_id=parse_str(data.get("variant_group_id")),
                variant_label=parse_str(data.get("variant_label")),
                launch_year=parse_int(data.get("launch_year")),
                os=parse_str(data.get("os")),
                current_price=parse_num(data.get("current_price")),
                weight_g=parse_int(data.get("weight_g")),
                height_mm=parse_num(data.get("height_mm")),
                width_mm=parse_num(data.get("width_mm")),
                thickness_mm=parse_num(data.get("thickness_mm")),
                overall_score=parse_num(data.get("overall_score")),
                image_url=parse_str(data.get("image_url")),
            )
            db.add(product)
            db.flush()

            # Display
            db.add(DisplaySpecs(
                product_id=product.id,
                screen_size_in=parse_num(data.get("screen_size_in")),
                screen_type=parse_str(data.get("screen_type")),
                resolution_width=parse_int(data.get("resolution_width")),
                resolution_height=parse_int(data.get("resolution_height")),
                refresh_rate_hz=parse_int(data.get("refresh_rate_hz")),
                brightness_nits=parse_int(data.get("brightness_nits")),
                ppi=parse_int(data.get("ppi")),
                protection=parse_str(data.get("protection")),
                hdr=parse_str(data.get("hdr")),
                aspect_ratio=parse_str(data.get("aspect_ratio")),
            ))

            # Hardware
            db.add(HardwareSpecs(
                product_id=product.id,
                chipset=parse_str(data.get("chipset")),
                cpu=parse_str(data.get("cpu")),
                gpu=parse_str(data.get("gpu")),
                ram_gb=parse_int(data.get("ram_gb")),
                storage_gb=parse_int(data.get("storage_gb")),
                battery_mah=parse_int(data.get("battery_mah")),
                charging_w=parse_int(data.get("charging_w")),
                charging_wireless_w=parse_int(data.get("charging_wireless_w")),
                charging_reverse=parse_bool(data.get("charging_reverse")),
            ))

            # Phone specs
            db.add(PhoneSpecs(
                product_id=product.id,
                sim_slots=parse_int(data.get("sim_slots")),
                has_esim=parse_bool(data.get("has_esim")),
                sim_type=parse_str(data.get("sim_type")),
                network_2g=parse_bool(data.get("network_2g")),
                network_3g=parse_bool(data.get("network_3g")),
                network_4g=parse_bool(data.get("network_4g")),
                network_5g=parse_bool(data.get("network_5g")),
                wifi=parse_str(data.get("wifi")),
                bluetooth=parse_str(data.get("bluetooth")),
                nfc=parse_bool(data.get("nfc")),
                usb_type=parse_str(data.get("usb_type")),
                usb_version=parse_str(data.get("usb_version")),
                ip_rating=parse_str(data.get("ip_rating")),
                has_memory_card=parse_bool(data.get("has_memory_card")),
                speakers=parse_str(data.get("speakers")),
                has_headphone_jack=parse_bool(data.get("has_headphone_jack")),
                has_fingerprint=parse_bool(data.get("has_fingerprint")),
            ))

            # Camera
            db.add(CameraSpecs(
                product_id=product.id,
                main_camera_mp=parse_num(data.get("main_camera_mp")),
                main_camera_aperture=parse_str(data.get("main_camera_aperture")),
                main_camera_ois=parse_bool(data.get("main_camera_ois")),
                camera_features=parse_str(data.get("camera_features")),
                ultrawide_mp=parse_num(data.get("ultrawide_mp")),
                ultrawide_aperture=parse_str(data.get("ultrawide_aperture")),
                telephoto_mp=parse_num(data.get("telephoto_mp")),
                telephoto_aperture=parse_str(data.get("telephoto_aperture")),
                telephoto_zoom=parse_str(data.get("telephoto_zoom")),
                video_resolution=parse_str(data.get("video_resolution")),
                front_camera_mp=parse_num(data.get("front_camera_mp")),
                front_camera_aperture=parse_str(data.get("front_camera_aperture")),
                front_video_resolution=parse_str(data.get("front_video_resolution")),
            ))

            db.commit()
            print(f"✅ Row {row_idx} — '{name}' imported successfully")
            success += 1

        except Exception as e:
            db.rollback()
            print(f"❌ Row {row_idx} — '{name}' error: {e}")
            errors += 1

    db.close()
    print(f"\n{'─'*40}")
    print(f"✅ Success: {success} products")
    print(f"❌ Errors:  {errors} products")
    print(f"{'─'*40}")


if __name__ == "__main__":
    filepath = os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        "product_import_template.xlsx"
    )

    if len(sys.argv) > 1:
        filepath = sys.argv[1]

    import_products(filepath)