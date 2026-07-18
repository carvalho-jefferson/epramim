import sys
import os
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_template():
    wb = Workbook()

    teal       = PatternFill("solid", fgColor="4EABB6")
    teal_light = PatternFill("solid", fgColor="E0F4F6")
    white      = PatternFill("solid", fgColor="FFFFFF")

    header_font  = Font(bold=True, color="FFFFFF", size=11)
    normal_font  = Font(size=10)
    example_font = Font(size=10, color="92400E", italic=True)

    thin_border = Border(
        left=Side(style="thin", color="E5E7EB"),
        right=Side(style="thin", color="E5E7EB"),
        top=Side(style="thin", color="E5E7EB"),
        bottom=Side(style="thin", color="E5E7EB"),
    )

    columns = [
        # GERAL
        ("name",                "Samsung Galaxy A55 8GB/256GB", 32, "GENERAL"),
        ("brand",               "Samsung",                      16, "GENERAL"),
        ("model",               "",                             16, "GENERAL"),
        ("slug",                "samsung-galaxy-a55-8gb-256gb", 32, "GENERAL"),
        ("variant_group_id",    "samsung-galaxy-a55",           24, "GENERAL"),
        ("variant_label",       "8GB/256GB",                    16, "GENERAL"),
        ("category",            "phone",                        12, "GENERAL"),
        ("launch_year",         "2024",                         14, "GENERAL"),
        ("os",                  "Android 14",                   18, "GENERAL"),
        ("current_price",       "1899.90",                      16, "GENERAL"),
        ("weight_g",            "213",                          12, "GENERAL"),
        ("height_mm",           "161.10",                       14, "GENERAL"),
        ("width_mm",            "77.40",                        14, "GENERAL"),
        ("thickness_mm",        "8.20",                         14, "GENERAL"),
        ("overall_score",       "",                             14, "GENERAL"),
        ("image_url",           "",                             30, "GENERAL"),
        # DISPLAY
        ("screen_size_in",      "6.60",                         16, "DISPLAY"),
        ("screen_type",         "Super AMOLED",                 18, "DISPLAY"),
        ("resolution_width",    "1080",                         18, "DISPLAY"),
        ("resolution_height",   "2340",                         18, "DISPLAY"),
        ("refresh_rate_hz",     "120",                          16, "DISPLAY"),
        ("brightness_nits",     "1000",                         16, "DISPLAY"),
        ("ppi",                 "390",                          10, "DISPLAY"),
        ("protection",          "Gorilla Glass Victus+",        24, "DISPLAY"),
        ("hdr",                 "HDR10+",                       14, "DISPLAY"),
        ("aspect_ratio",        "19.5:9",                       14, "DISPLAY"),
        # HARDWARE
        ("chipset",             "Exynos 1480",                  20, "HARDWARE"),
        ("cpu",                 "Octa-core",                    18, "HARDWARE"),
        ("gpu",                 "Xclipse 530",                  18, "HARDWARE"),
        ("ram_gb",              "8",                            12, "HARDWARE"),
        ("storage_gb",          "256",                          14, "HARDWARE"),
        ("battery_mah",         "5000",                         14, "HARDWARE"),
        ("charging_w",          "25",                           14, "HARDWARE"),
        ("charging_wireless_w", "",                             18, "HARDWARE"),
        ("charging_reverse",    "FALSE",                        18, "HARDWARE"),
        # PHONE
        ("sim_slots",           "2",                            12, "PHONE"),
        ("has_esim",            "TRUE",                         12, "PHONE"),
        ("sim_type",            "Nano SIM",                     14, "PHONE"),
        ("network_2g",          "TRUE",                         12, "PHONE"),
        ("network_3g",          "TRUE",                         12, "PHONE"),
        ("network_4g",          "TRUE",                         12, "PHONE"),
        ("network_5g",          "TRUE",                         12, "PHONE"),
        ("wifi",                "Wi-Fi 6",                      14, "PHONE"),
        ("bluetooth",           "5.3",                          14, "PHONE"),
        ("nfc",                 "TRUE",                         10, "PHONE"),
        ("usb_type",            "USB-C",                        12, "PHONE"),
        ("usb_version",         "2.0",                          12, "PHONE"),
        ("ip_rating",           "IP67",                         12, "PHONE"),
        ("has_memory_card",     "FALSE",                        16, "PHONE"),
        ("speakers",            "Stereo",                       14, "PHONE"),
        ("has_headphone_jack",  "FALSE",                        20, "PHONE"),
        ("has_fingerprint",     "TRUE",                         16, "PHONE"),
        # CAMERA
        ("main_camera_mp",      "50",                           16, "CAMERA"),
        ("main_camera_aperture","f/1.8",                        20, "CAMERA"),
        ("main_camera_ois",     "TRUE",                         16, "CAMERA"),
        ("camera_features",     "PDAF, OIS, HDR10+",            24, "CAMERA"),
        ("ultrawide_mp",        "12",                           14, "CAMERA"),
        ("ultrawide_aperture",  "f/2.2",                        18, "CAMERA"),
        ("telephoto_mp",        "",                             14, "CAMERA"),
        ("telephoto_aperture",  "",                             18, "CAMERA"),
        ("telephoto_zoom",      "",                             16, "CAMERA"),
        ("video_resolution",    "4K @30fps",                    18, "CAMERA"),
        ("front_camera_mp",     "32",                           16, "CAMERA"),
        ("front_camera_aperture","f/2.2",                       20, "CAMERA"),
        ("front_video_resolution","4K @30fps",                  20, "CAMERA"),
    ]

    section_colors = {
        "GENERAL":  ("4EABB6", "E0F4F6"),
        "DISPLAY":  ("2D6A9F", "DBEAFE"),
        "HARDWARE": ("7C3AED", "EDE9FE"),
        "PHONE":    ("059669", "D1FAE5"),
        "CAMERA":   ("D97706", "FEF3C7"),
    }

    ws = wb.active
    ws.title = "Products"
    ws.freeze_panes = "A4"

    # Linha 1 — seções agrupadas
    current_section = None
    section_start = 1
    for i, (col_name, _, width, section) in enumerate(columns, 1):
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = width

        if section != current_section:
            if current_section is not None:
                ws.merge_cells(
                    start_row=1, start_column=section_start,
                    end_row=1, end_column=i - 1
                )
            current_section = section
            section_start = i

        bg, _ = section_colors[section]
        cell = ws.cell(row=1, column=i, value=section if i == section_start else "")
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells(
        start_row=1, start_column=section_start,
        end_row=1, end_column=len(columns)
    )
    ws.row_dimensions[1].height = 22

    # Linha 2 — nomes das colunas
    for i, (col_name, _, _, section) in enumerate(columns, 1):
        _, light = section_colors[section]
        cell = ws.cell(row=2, column=i, value=col_name)
        cell.fill = PatternFill("solid", fgColor=light)
        cell.font = Font(bold=True, color="1A1A1A", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    ws.row_dimensions[2].height = 30

    # Linha 3 — exemplos
    for i, (_, example, _, section) in enumerate(columns, 1):
        cell = ws.cell(row=3, column=i, value=example)
        cell.fill = PatternFill("solid", fgColor="FFFBEB")
        cell.font = Font(size=10, color="92400E", italic=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    ws.row_dimensions[3].height = 18

    # Linhas de dados
    for row in range(4, 104):
        for col in range(1, len(columns) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = white if row % 2 == 0 else PatternFill("solid", fgColor="F9FAFB")
            cell.font = normal_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
        ws.row_dimensions[row].height = 18

    # Validação de slug duplicado — destaca vermelho se repetido
    from openpyxl.formatting.rule import FormulaRule
    slug_col = next(i for i, (h, _, _, _) in enumerate(columns, 1) if h == "slug")
    slug_letter = get_column_letter(slug_col)
    red_fill = PatternFill("solid", fgColor="FECACA")
    ws.conditional_formatting.add(
        f"{slug_letter}4:{slug_letter}103",
        FormulaRule(
            formula=[f'COUNTIF(${slug_letter}$4:${slug_letter}$103,{slug_letter}4)>1'],
            fill=red_fill
        )
    )

    # Aba de instruções
    wi = wb.create_sheet("Instructions")
    wi.column_dimensions["A"].width = 80

    instructions = [
        ("PRODUCT IMPORT TEMPLATE — INSTRUCTIONS", True, "4EABB6", "FFFFFF"),
        ("", False, "FFFFFF", "000000"),
        ("HOW TO USE", True, "E0F4F6", "1A6570"),
        ("1. Fill data starting from row 4 (row 3 is example only).", False, "FFFFFF", "1A1A1A"),
        ("2. Each row = one product.", False, "FFFFFF", "1A1A1A"),
        ("3. Required fields: name, brand, slug, category", False, "FFFFFF", "1A1A1A"),
        ("4. TRUE/FALSE fields: use TRUE or FALSE (case insensitive)", False, "FFFFFF", "1A1A1A"),
        ("5. Leave empty cells blank — don't use 0 for missing values", False, "FFFFFF", "1A1A1A"),
        ("6. Duplicate slugs highlight RED automatically", False, "FFFFFF", "1A1A1A"),
        ("7. Run: python scripts/import_from_excel.py", False, "FFFFFF", "1A1A1A"),
        ("", False, "FFFFFF", "000000"),
        ("VARIANTS", True, "E0F4F6", "1A6570"),
        ("Same product, different RAM/storage = separate rows with same variant_group_id", False, "FFFFFF", "1A1A1A"),
        ("Example: samsung-galaxy-a55-8gb-256gb and samsung-galaxy-a55-8gb-128gb", False, "FFFFFF", "1A1A1A"),
        ("variant_group_id: samsung-galaxy-a55 (same for both)", False, "FFFFFF", "1A1A1A"),
        ("variant_label: 8GB/256GB and 8GB/128GB", False, "FFFFFF", "1A1A1A"),
        ("", False, "FFFFFF", "000000"),
        ("VIDEO RESOLUTION FORMAT", True, "E0F4F6", "1A6570"),
        ("Use format: 4K @30fps or 4K @30/60fps", False, "FFFFFF", "1A1A1A"),
        ("Same format for front_video_resolution", False, "FFFFFF", "1A1A1A"),
        ("", False, "FFFFFF", "000000"),
        ("SLUG FORMAT", True, "E0F4F6", "1A6570"),
        ("✓ samsung-galaxy-a55-8gb-256gb", False, "FFFFFF", "059669"),
        ("✓ apple-iphone-16-pro-256gb", False, "FFFFFF", "059669"),
        ("✗ Samsung Galaxy A55 (no spaces)", False, "FFFFFF", "DC2626"),
        ("✗ samsung_galaxy_a55 (no underscores)", False, "FFFFFF", "DC2626"),
    ]

    for row_idx, (text, bold, bg, fg) in enumerate(instructions, 1):
        cell = wi.cell(row=row_idx, column=1, value=text)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.font = Font(bold=bold, color=fg, size=11 if bold else 10)
        cell.alignment = Alignment(vertical="center", indent=1)
        wi.row_dimensions[row_idx].height = 22 if bold else 18

    output_path = os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        "product_import_template.xlsx"
    )
    wb.save(output_path)
    print(f"✅ Template created: {output_path}")


if __name__ == "__main__":
    create_template()